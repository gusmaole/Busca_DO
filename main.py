from selenium.webdriver import Chrome
from time import sleep
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

#Acessar URL do Diário oficial
browser = Chrome()
url = 'http://www.imprensaoficial.com.br/'
browser.get(url)

#Clicar em busca avançada
busca_avancada = browser.find_element(by=By.XPATH, value='//*[@id="buscaporpalavra"]/a')
busca_avancada.click()
sleep(2)

#Digitar a palavra chave
palavra_chave = browser.find_element(by=By.ID, value='txtPalavrasChave')
palavra_chave.send_keys('"Leandro Gusmão da Silva"')

sleep(1)
#Todos os cadernos
check_box = browser.find_element(by=By.XPATH, value='//*[@id="content_content_content_chkGrupos_0"]')
check_box.click()

#Buscar
buscar = browser.find_element(by=By.XPATH, value='//*[@id="content_content_content_btnBuscar"]')
buscar.click()

sleep(4)

dicas = browser.find_element(by=By.XPATH, value='/html/body/div[1]/div/a[2]')
dicas.click()
sleep(1)

#Ordenar por data
ordenar_data = browser.find_element(by=By.XPATH, value='//*[@id="content_lnkOrderByData"]')
ordenar_data.click()
sleep(2)

#Copiar título
titulo = browser.find_elements(by=By.XPATH, value='//*[@id="form"]/div[3]/div/div[2]/div/div[1]')

#Copiar trecho
trecho = browser.find_elements(by=By.XPATH, value='//*[@id="form"]/div[3]/div/div[2]/div/div[2]')

sleep(2)

#Importar no Excel
wb = Workbook()
ws = wb.active

#Cabeçalhos
ws['A1'] = "Título"
ws['B1'] = "Trecho"

#Adicionando os dados na planilha
for index, elemento in enumerate(titulo, start=2):
    valor = elemento.text  # Extraia o texto do elemento Selenium
    ws.cell(row=index, column=1, value=valor)  # Insira na coluna A

for index, elemento in enumerate(trecho, start=2):
    valor = elemento.text  # Extraia o texto do elemento Selenium
    ws.cell(row=index, column=2, value=valor)  # Insira na coluna B

#Salva o arquivo
wb.save("DO.xlsx")

#Enviar por e-mail

# Configurações do servidor SMTP
smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_username = '' #E-mail que irá enviar a mensagem
smtp_password = '' #Senha

# Crie um objeto MIMEMultipart
msg = MIMEMultipart()

# Configuração dos campos do e-mail
msg['From'] = smtp_username
msg['To'] = '' #Inserir e-mail destino
msg['Subject'] = 'Atualização diário oficial' #Título do e-mail

# Corpo do e-mail
corpo = 'Segue anexo atualização semanal do diário oficial'
msg.attach(MIMEText(corpo, 'plain'))

# Anexar um arquivo ao e-mail
arquivo_anexo = 'DO.xlsx'
with open(arquivo_anexo, 'rb') as f:
    anexo = MIMEApplication(f.read(), Name='DO.xlsx')
anexo['Content-Disposition'] = f'attachment; filename="{arquivo_anexo}"'
msg.attach(anexo)

# Conecte-se ao servidor SMTP
server = smtplib.SMTP(smtp_server, smtp_port)
server.starttls()
server.login(smtp_username, smtp_password)

# Envie o e-mail
server.sendmail(smtp_username, 'leandrogusmao@professor.educacao.sp.gov.br', msg.as_string())

# Encerre a conexão com o servidor SMTP
server.quit()
